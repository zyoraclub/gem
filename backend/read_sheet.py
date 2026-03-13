from openpyxl import load_workbook

wb = load_workbook('/Users/redfoxhotels/gem/sheet/Copy of Magnesium Sulphate (Epsom Salt) as per IS 2730.xlsx')
ws = wb.active

print('Sheet:', ws.title)
print('Rows:', ws.max_row, '| Cols:', ws.max_column)
print()
print('=== HEADERS ===')
for i, cell in enumerate(ws[1], 1):
    letter = chr(64+i) if i <= 26 else 'X'
    print(f'{letter}: {cell.value}')
print()
print('=== ROW 2 SAMPLE ===')
for i, cell in enumerate(ws[2], 1):
    letter = chr(64+i) if i <= 26 else 'X'
    if cell.value:
        print(f'{letter}: {str(cell.value)[:60]}')
